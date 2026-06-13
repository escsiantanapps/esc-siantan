"""
Script migrasi data dari Excel ke Supabase.
Jalankan sekali saja untuk import 193 jemaat existing.

Install dulu:
  pip install openpyxl supabase python-dotenv

Cara pakai:
  1. Letakkan file Database_CRM.xlsx di folder yang sama
  2. Buat file .env dengan SUPABASE_URL dan SUPABASE_SERVICE_KEY
  3. python migrate_data.py
"""

import os
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from supabase import create_client
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_KEY = os.environ["SUPABASE_SERVICE_KEY"]  # pakai service key, bukan anon key

supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

def excel_serial_to_date(serial):
    """Konversi serial Excel ke tanggal"""
    if not serial or not isinstance(serial, (int, float)):
        return None
    try:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(serial))).date().isoformat()
    except:
        return None

def clean_phone(phone):
    """Normalkan nomor telepon"""
    if not phone:
        return None
    s = str(phone).replace('.', '').replace(' ', '').replace('-', '')
    # Hapus desimal dari scientific notation
    if 'E' in s.upper():
        s = str(int(float(s)))
    if s.startswith('0'):
        return '+62' + s[1:]
    if s.startswith('62'):
        return '+' + s
    return s

def migrate_users():
    wb = load_workbook("Database_CRM.xlsx", read_only=True)
    ws = wb["Users"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"Ditemukan {len(rows)} baris data jemaat")

    success, skipped = 0, 0

    for row in rows:
        if not row[0] or not row[3]:  # skip baris kosong
            continue

        user_id, username, _, name, role, status, ministry_str, email, \
        birth_serial, birth_place, photo_url, blood_type, gender, \
        social_media, phone, address, _, nik, komsel_id, sp_level, sp_notes = (list(row) + [None]*22)[:22]

        # Parse ministry IDs
        ministry_ids = []
        if ministry_str:
            ministry_ids = [m.strip() for m in str(ministry_str).split(',') if m.strip()]

        record = {
            "user_id":      str(user_id) if user_id else None,
            "username":     str(username) if username else None,
            "name":         str(name).strip(),
            "email":        str(email).strip() if email and str(email) != 'nan' else None,
            "phone":        clean_phone(phone),
            "role":         str(role) if role in ['Volunteer','Jemaat','Admin','Super Admin','PKS'] else 'Jemaat',
            "status":       str(status) if status in ['Aktif','Nonaktif'] else 'Aktif',
            "gender":       str(gender) if gender in ['Laki-laki','Perempuan'] else None,
            "birth_date":   excel_serial_to_date(birth_serial),
            "birth_place":  str(birth_place) if birth_place and str(birth_place) not in ['-','nan'] else None,
            "address":      str(address) if address and str(address) not in ['-','nan'] else None,
            "blood_type":   str(blood_type) if blood_type and str(blood_type) not in ['-','nan'] else None,
            "social_media": str(social_media) if social_media and str(social_media) not in ['-','nan'] else None,
            "photo_url":    str(photo_url) if photo_url else None,
            "nik":          str(nik) if nik else None,
            "ministry_ids": ministry_ids,
            "komsel_id":    str(komsel_id) if komsel_id else None,
            "sp_level":     str(sp_level) if sp_level in ['Aman','SP 1','SP 2','SP 3'] else 'Aman',
            "sp_notes":     str(sp_notes) if sp_notes and str(sp_notes) not in ['-','nan'] else None,
        }

        try:
            supabase.table("users").upsert(record, on_conflict="user_id").execute()
            success += 1
            print(f"  ✓ {name}")
        except Exception as e:
            skipped += 1
            print(f"  ✗ {name}: {e}")

    wb.close()
    print(f"\nSelesai: {success} berhasil, {skipped} dilewati")

def migrate_ministries():
    wb = load_workbook("Database_CRM.xlsx", read_only=True)
    ws = wb["MinistryData"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    success = 0

    for row in rows:
        if not row[0]: continue
        ministry_id, name = row[0], row[1]
        try:
            supabase.table("ministries").upsert(
                {"ministry_id": str(ministry_id), "name": str(name)},
                on_conflict="ministry_id"
            ).execute()
            success += 1
        except Exception as e:
            print(f"  ✗ {name}: {e}")

    wb.close()
    print(f"Ministry: {success} berhasil diimport")

def migrate_komsel():
    wb = load_workbook("Database_CRM.xlsx", read_only=True)
    ws = wb["KomselData"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    success = 0

    for row in rows:
        if not row[0]: continue
        komsel_id, name, leader, capacity = (list(row) + [None, None])[:4]
        try:
            supabase.table("komsel").upsert({
                "komsel_id":    str(komsel_id),
                "name":         str(name),
                "leader_name":  str(leader) if leader else None,
                "max_capacity": int(capacity) if capacity else 20,
            }, on_conflict="komsel_id").execute()
            success += 1
        except Exception as e:
            print(f"  ✗ {name}: {e}")

    wb.close()
    print(f"Komsel: {success} berhasil diimport")

if __name__ == "__main__":
    print("=== Migrasi Data GerejaKu ===\n")
    print("1. Import Ministry...")
    migrate_ministries()
    print("\n2. Import Komsel...")
    migrate_komsel()
    print("\n3. Import Users/Jemaat...")
    migrate_users()
    print("\n=== Migrasi selesai! ===")
    print("PENTING: Reset semua password jemaat via Supabase Auth,")
    print("karena password lama tidak bisa dipakai (plain text).")
